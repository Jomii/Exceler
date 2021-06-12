import locale
locale.setlocale(locale.LC_ALL, "fi_FI")
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from sys import argv
import re
from collections import defaultdict
import datetime

# Start program:
# python Exceler.py inputfilename.txt

# TODO: Add errorchecking for filename input

income = []   # A list of (date, amount, payer) tuples.
expenses = [] # A list of (date, amount, receiver) tuples.

if len(argv) != 3:
    print("usage: python Exceler.py <excel file> <transactions file>")
    exit()

fileName = argv[1]
transactionFilename = argv[2]

with open(transactionFilename) as f:
    # Skip header of data file with next(f).
    next(f)
    next(f)
    next(f)
    for line in f:
        parts = line.strip('\n').split('\t')

        # Skip empty parts that splitting with \t sometimes causes.
        if len(parts) > 1:
            # Create a tuple with the data we want.
            t = parts[2], float(parts[3].replace(',', '.')), parts[4]

            if t[1] > 0:
                income.append(t)
            else:
                expenses.append(t)


print(f"Adding data from {transactionFilename} to {fileName}")
wb = openpyxl.load_workbook(fileName)
# Get month of first entry
datetime_obj = datetime.datetime.strptime(str(income[0][0]), "%x")
month = datetime_obj.strftime("%B")
year = datetime_obj.strftime("%y")
print(f"Creating new work sheet with name '{month + year}'")
ws = wb.create_sheet(month + year, -1)

""" Add income data to cells. """
ws["A1"] = "Pvm."
ws["B1"] = "€"
ws["C1"] = "Saaja/Maksaja"

i = 2

for day in income:
    ws[f"A{i}"] = day[0].split('.')[0]
    ws[f"B{i}"] = day[1]
    ws[f"C{i}"] = day[2]

    i = i + 1

i = i + 2 # Separate values by two rows.

""" Add expense data to cells. """
ws[f"A{i - 1}"] = "Pvm."
ws[f"B{i - 1}"] = "€"
ws[f"C{i - 1}"] = "Saaja/Maksaja"

for day in expenses:
    ws[f"A{i}"] = day[0].split('.')[0]
    ws[f"B{i}"] = day[1]
    ws[f"C{i}"] = day[2]

    i = i + 1


ws.insert_cols(1, amount=1) # Add a column above the data in the sheet.
ws.insert_rows(1, amount=1) # Add a row on the left of the data.
ws.column_dimensions["C"].width = 10 # Set the width of date column.
ws.column_dimensions["D"].width = 35 # Set the width of payer/receiver column.

# Vertical text in the first column for income and expenses.
ws["A1"] = "TULOT"
ws[f"A{len(income) + 3}"] = "MENOT"

incomeC = ws["E1"] # Cell where the total income is displayed.
expensesC = ws[f"E{len(income) + 3}"] # Cell where the total expenses are displayed.

# Add the balance of the month to a cell.
g1 = ws["G1"]
h1 = ws["H1"]
g1.value = "Balanssi:"
h1.value = f"=SUM(E1:E{len(income) + 3})"

# Apply styles to a range of cells as if they were a single cell.
def style_range(ws, cell_range, fill=None, font=None, alignment=None, merge=False):
    # If param. merge is true, then merge cells in range and apply style to them.
    if merge:
        first_cell = ws[cell_range.split(":")[0]]
        ws.merge_cells(cell_range)
        first_cell.fill = fill
        first_cell.alignment = alignment
        first_cell.font = font
    # Apply styles to each cell individually.
    else:
        rows = ws[cell_range]
        for row in rows:
            for c in row:
                if fill:
                    c.fill = fill
                if font:
                    c.font = font
                if alignment:
                    c.alignment = alignment


# Apply styling to the cells.
greenColor = PatternFill(fill_type='solid', start_color='c6e0b4', end_color='c6e0b4')
redColor = PatternFill(fill_type='solid', start_color='fce4d6', end_color='fce4d6')
yellowColor = PatternFill(fill_type='solid', start_color='ffe699', end_color='ffe699')
centerAlign = Alignment(horizontal="center", vertical="center")
centerHorizontal = Alignment(horizontal="center", vertical="center", text_rotation=90)
bold = Font(bold=True)
double = Side(border_style="medium")
border = Border(top=double, left=double, right=double, bottom=double)
# Add colors and cell alignments
style_range(ws, 'A1:D2', fill=greenColor)
style_range(ws, f'A{len(income) + 3}:D{len(income) + 4}', fill=redColor)
style_range(ws, f'A1:A{len(income) + 2}', fill=greenColor, font=bold, alignment=centerHorizontal, merge=True)
style_range(ws, f'A{len(income) + 3}:A{i}', fill=redColor, font=bold, alignment=centerHorizontal, merge=True)
# For balance
style_range(ws, 'G1:H1', font=bold, fill=yellowColor)
# Add fonts
style_range(ws, 'B2:D2', font=bold)
style_range(ws, f'B{len(income) + 4}:D{len(income) + 4}', font=bold)
# Center Date and currency
style_range(ws, f'B2:C{i}', alignment=centerAlign)
# Merge, center and add borders for totals
style_range(ws, 'E1:E2', fill=greenColor, font=bold, alignment=centerAlign, merge=True)
ws["E1"].border = border
ws["E2"].border = border
style_range(ws, f'E{len(income) + 3}:E{len(income) + 4}', fill=redColor, font=bold, alignment=centerAlign, merge=True)
ws[f"E{len(income) + 3}"].border = border
ws[f"E{len(income) + 4}"].border = border
incomeC.value = sum(map(lambda x: x[1], income))
expensesC.value = sum(map(lambda x: x[1], expenses))

wb.save(fileName) # Save changes to the workbook.